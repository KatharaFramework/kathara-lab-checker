import os

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet


def write_final_results_to_excel(test_collector: "TestCollectorPackage.TestCollector", path: str):
    # Create a new Excel workbook
    workbook = Workbook()

    # Select the active sheet
    sheet = workbook.active

    sheet["A1"] = "Student Name"
    sheet["B1"] = "Tests Passed"
    sheet["C1"] = "Tests Failed"
    sheet["D1"] = "Tests Total Number"
    sheet["E1"] = "Problems"

    for index, (test_name, test_results) in enumerate(test_collector.tests.items()):
        failed_tests = test_collector.get_failed(test_name)
        passed_tests = test_collector.get_passed(test_name)
        sheet["A" + str(index + 2)] = test_name
        sheet["B" + str(index + 2)] = len(passed_tests)
        sheet["C" + str(index + 2)] = len(failed_tests)
        sheet["D" + str(index + 2)] = len(test_results)

        if failed_tests:
            failed_string = ""
            for idx, failed in enumerate(failed_tests):
                failed_string += f"{(idx + 1)}: {failed.description}: {failed.reason}\n"
            if len(failed_string) >= 32767:
                raise Exception("ERROR: Excel cell too big")
            sheet["E" + str(index + 2)] = failed_string
            sheet["E" + str(index + 2)].alignment = Alignment(wrapText=True)
        else:
            sheet["E" + str(index + 2)] = "None"

    excel_file = os.path.join(path, "results.xlsx")
    workbook.save(excel_file)


def _write_sheet_row(sheet: Worksheet, column: int, description: str, passed: str, reason: str) -> None:
    sheet["A" + str(column + 2)] = description
    sheet["B" + str(column + 2)] = passed
    sheet["C" + str(column + 2)] = reason


def write_result_to_excel(check_results: list["CheckResultPackage.CheckResult"], path: str):
    # Create a new Excel workbook
    workbook: Workbook = Workbook()

    workbook.create_sheet("Summary", 0)
    sheet_summary = workbook.get_sheet_by_name("Summary")
    sheet_summary["A1"] = "Total Tests"
    sheet_summary["B1"] = "Passed Tests"
    sheet_summary["C1"] = "Failed"

    _write_sheet_row(
        sheet_summary,
        0,
        str(len(check_results)),
        str(len(list(filter(lambda x: x.passed, check_results)))),
        str(len(list(filter(lambda x: not x.passed, check_results)))),
    )

    # Select the active sheet
    workbook.create_sheet("All", 1)
    sheet_all = workbook.get_sheet_by_name("All")
    sheet_all["A1"] = "Tests Description"
    sheet_all["B1"] = "Passed"
    sheet_all["C1"] = "Reason"

    workbook.create_sheet("Failed", 2)
    sheet_failed = workbook.get_sheet_by_name("Failed")
    sheet_failed["A1"] = "Tests Description"
    sheet_failed["B1"] = "Passed"
    sheet_failed["C1"] = "Reason"

    failed_index = 0
    for index, check_result in enumerate(check_results):
        if not check_result.passed:
            _write_sheet_row(
                sheet_failed,
                failed_index,
                check_result.description,
                str(check_result.passed),
                check_result.reason,
            )
            failed_index += 1
        _write_sheet_row(sheet_all, index, check_result.description, str(check_result.passed), check_result.reason)
    workbook.save(os.path.join(path, f"{os.path.basename(path)}_result.xlsx"))
